import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors

from collections import Iterable

wb = Workbook()
sheet = wb.active
sheet.title = "new sheet"
sheet['C3'] = 'hello'


#bold_itatic24_font = Font(name='等线', size=24, italic=True, color = colors.RED, bold=True)
#sheet['A1'].font = bold_itatic24_font
# 设置B1中的数据垂直居中和水平居中
sheet['B1'].alignment = Alignment(horizontal = 'center', vertical = 'center')
# 第2行行高
sheet.row_dimensions[2].height = 40
# C列列宽
sheet.column_dimensions['C'].width = 40

for i in range(10):
    sheet["A%d" % (i+1)].value = i+1
sheet["E1"].value = "=SUM(A:A)"

# sheet.merge_cells('B1:G1') # 合并一行中的几个单元格
# sheet.merge_cells('A1:C3') # 合并一个矩形区域中的单元格



###### Standard format ########
#######       合并    #########
def format_init(sheet):
    A1_fill = PatternFill(fill_type = 'solid', start_color = colors.YELLOW, end_color = colors.BLUE)
    A1_font = Font(name = '微软雅黑', size = 14, bold = True, color = 'FF000000')
    A1_alignment = Alignment(horizontal = 'center', vertical = 'center')
    A1_border = Border(left = Side(border_style = 'thin', color = colors.BLACK),
                       right = Side(border_style = 'thin', color = colors.BLACK),
                       top = Side(border_style = 'thin', color = colors.BLACK),
                       bottom = Side(border_style = 'thin', color = colors.BLACK))
    sheet['A22'].font = A1_font
    sheet['A22'].fill = A1_fill
    sheet['A22'].alignment = A1_alignment
    sheet['A22'].border = A1_border
    sheet.merge_cells('A22:I22')  #target cell   A1
    sheet.row_dimensions[22].height = 40
    sheet['A22'] = 'huayuans'
    for objsheetcells in sheet['A22:I22']:
        for cellone in objsheetcells:
            cellone.border = A1_border
    #for x in 'A:I':
        # sheet[.border = A1_border

def format_second_line(sheet):
    A1_fill = PatternFill(fill_type = 'solid', start_color = colors.YELLOW)
    A1_font = Font(name = '微软雅黑', size = 12, bold = True, color = 'FF000000')
    A1_alignment = Alignment(horizontal = 'center', vertical = 'center')
    A1_border = Border(left = Side(border_style = 'thin', color = colors.BLACK),
                       right = Side(border_style = 'thin', color = colors.BLACK),
                       top = Side(border_style = 'thin', color = colors.BLACK),
                       bottom = Side(border_style = 'thin', color = colors.BLACK), )
    for objsheetcells in sheet['A33:A36']:
        for cellone in objsheetcells:
            cellone.value = 1
            cellone.border = A1_border


# set cells border
def set_border(sheet, cell_scope):
    _border = Border(left = Side(border_style = 'thin', color = colors.BLACK),
                   right = Side(border_style = 'thin', color = colors.BLACK),
                   top = Side(border_style = 'thin', color = colors.BLACK),
                   bottom = Side(border_style = 'thin', color = colors.BLACK))
    # if there is only one cell, no iteration is executed.
    if isinstance(sheet[cell_scope],Iterable) is not True:
        sheet[cell_scope].border = _border
    else:
        for _cells in sheet[cell_scope]:
            for _cell in _cells:
                _cell.border = _border


format_init(sheet)
set_border(sheet,'A36')


wb.save("data/saveNew.xlsx")




