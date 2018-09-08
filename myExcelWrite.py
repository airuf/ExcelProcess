import openpyxl
import myExcelRead
import operator
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors

from collections import Iterable




# sheet.title = "new sheet"
# sheet['C3'] = 'hello'
#
#
# #bold_itatic24_font = Font(name='等线', size=24, italic=True, color = colors.RED, bold=True)
# #sheet['A1'].font = bold_itatic24_font
# # 设置B1中的数据垂直居中和水平居中
# sheet['B1'].alignment = Alignment(horizontal = 'center', vertical = 'center')
# # 第2行行高
# sheet.row_dimensions[2].height = 40
# # C列列宽
# sheet.column_dimensions['C'].width = 40
#
# for i in range(10):
#     sheet["A%d" % (i+1)].value = i+1
# sheet["E1"].value = "=SUM(A:A)"

# sheet.merge_cells('B1:G1') # 合并一行中的几个单元格
# sheet.merge_cells('A1:C3') # 合并一个矩形区域中的单元格



###### Standard format ########
#######       合并    #########
def format_init(sheet):
    A1_fill = PatternFill(fill_type = 'solid', start_color = colors.YELLOW, end_color = colors.BLUE)
    A1_font = Font(name = '微软雅黑', size = 14, bold = True, color = 'FF000000')
    A1_alignment = Alignment(horizontal = 'center', vertical = 'center')
    A1_border = Border(left = Side(border_style = 'thin', color = colors.BLACK),
                       right = Side(border_style = 'thin', color = colors.BLACK),
                       top = Side(border_style = 'thin', color = colors.BLACK),
                       bottom = Side(border_style = 'thin', color = colors.BLACK))
    sheet['A22'].font = A1_font
    sheet['A22'].fill = A1_fill
    sheet['A22'].alignment = A1_alignment
    sheet['A22'].border = A1_border
    sheet.merge_cells('A22:I22')  #target cell   A1
    sheet.row_dimensions[22].height = 40
    sheet['A22'] = 'huayuans'
    for objsheetcells in sheet['A22:I22']:
        for cellone in objsheetcells:
            cellone.border = A1_border
    #for x in 'A:I':
        # sheet[.border = A1_border

def format_second_line(sheet):
    A1_fill = PatternFill(fill_type = 'solid', start_color = colors.YELLOW)
    A1_font = Font(name = '微软雅黑', size = 12, bold = True, color = 'FF000000')
    A1_alignment = Alignment(horizontal = 'center', vertical = 'center')
    A1_border = Border(left = Side(border_style = 'thin', color = colors.BLACK),
                       right = Side(border_style = 'thin', color = colors.BLACK),
                       top = Side(border_style = 'thin', color = colors.BLACK),
                       bottom = Side(border_style = 'thin', color = colors.BLACK), )
    for objsheetcells in sheet['A33:A36']:
        for cellone in objsheetcells:
            cellone.value = 1
            cellone.border = A1_border


# set cells border
def set_border(sheet, cell_scope):
    _border = Border(left = Side(border_style = 'thin', color = colors.BLACK),
                   right = Side(border_style = 'thin', color = colors.BLACK),
                   top = Side(border_style = 'thin', color = colors.BLACK),
                   bottom = Side(border_style = 'thin', color = colors.BLACK))
    # if there is only one cell, no iteration is executed.
    if isinstance(sheet[cell_scope],Iterable) is not True:
        sheet[cell_scope].border = _border
    else:
        for _cells in sheet[cell_scope]:
            for _cell in _cells:
                _cell.border = _border


# format_init(sheet)
# set_border(sheet,'A36')



#set company title
def title_init(sheet, company_name):
    A1_fill = PatternFill(fill_type = 'solid', start_color = colors.YELLOW, end_color = colors.BLUE)
    A1_font = Font(name = '微软雅黑', size = 14, bold = True, color = 'FF000000')
    A1_alignment = Alignment(horizontal = 'center', vertical = 'center')
    A1_border = Border(left = Side(border_style = 'thick', color = colors.BLACK),
                       right = Side(border_style = 'thick', color = colors.BLACK),
                       top = Side(border_style = 'thick', color = colors.BLACK),
                       bottom = Side(border_style = 'thick', color = colors.BLACK))
    sheet['A1'].font = A1_font
    sheet['A1'].fill = A1_fill
    sheet['A1'].alignment = A1_alignment
    sheet['A1'].border = A1_border
    sheet.merge_cells('A1:I1')  #target cell   A1
    sheet.row_dimensions[1].height = 40
    sheet['A1'] = company_name
    for objsheetcells in sheet['A1:I1']:
        for cellone in objsheetcells:
            cellone.border = A1_border


def list_detail(sheet, details_list):
    B_fill = PatternFill(fill_type = 'solid', start_color = colors.YELLOW, end_color = colors.BLUE)
    B_font = Font(name = '微软雅黑', size = 11, bold = True, color = 'FF000000')
    B_alignment = Alignment(horizontal = 'center', vertical = 'center')
    B_border = Border(left = Side(border_style = 'thick', color = colors.BLACK),
                       right = Side(border_style = 'thick', color = colors.BLACK),
                       top = Side(border_style = 'thick', color = colors.BLACK),
                       bottom = Side(border_style = 'thick', color = colors.BLACK))
    len_column = details_list.__len__()
    for x in range(len_column):
        addr = chr(65 + x) + '2'
        sheet[addr] = details_list[x]
        sheet.column_dimensions[chr(65 + x)].width = details_list[x].__len__()*2.3
        sheet[addr].font = B_font
        sheet[addr].fill = B_fill
        sheet[addr].alignment = B_alignment
        sheet[addr].border = B_border

    # for x in range(len_column):
    #     sheet.row_values(2)[x] =1;

def total_set(sheet, range, data):
    sheet.merge_cells(range)  # target cell   A1
    addr = range.split(':')
    sheet[addr[0]] = data


def creat_target_file(StoreInfo):
    _StoreInfo = myExcelRead.MasterAccount(' ')
    _StoreInfo = StoreInfo
    wb = Workbook()
    sheet = wb.active
    title_init(sheet, '花苑')
    list = ['排名', '门店名称', '经办人', '成交单数', '商家服务费', '店员服务费', '金蛋金额', '店员收入合计', '门店服务费合计']
    list_detail(sheet, list)
    #total_set(sheet, 'A5:C5', 22)
    #

    # print message of StoreInfo
    x = myExcelRead.StoreInfo(' ')
    store_position = 3;

    #Form ordering
    ordering_list = []
    for x in _StoreInfo.store.values():
        x.total_result()
    # for x in _StoreInfo.store.items():
        # print("obj", x, x[1], type(x))
        # print("all income ", x[1].all_income)
    _StoreInfo = sorted(_StoreInfo.store.items(), key = lambda store:store[1].all_income, reverse=True)

    for m in _StoreInfo:
        x = m[1]
        b = x
        b.get_num_clerks()
        # print("numbers:", b.num_of_clerks)
        #TODO:set the number of rank to this merge cell
        addr = str('A' + str(store_position) + ':' + 'A' + str(b.num_of_clerks+store_position-1))
        total_set(sheet, addr, b.rank)

        addr = str('B' + str(store_position) + ':' + 'B' + str(b.num_of_clerks + store_position - 1))
        total_set(sheet, addr, b.name)
        addr = str('I' + str(store_position) + ':' + 'I' + str(b.num_of_clerks + store_position - 1))
        total_set(sheet, addr, b.all_income)
        #print(b.name)

        num = 0
        listMe = []
        for meme in b.clerks.values():
            listMe.append(meme)
            #TODO:
        for addr in range(store_position, b.num_of_clerks+store_position):
            sheet[str('C'+addr)] = str(listMe[num].name)
            num = num+1

        # print(b.name, "income", b.all_income)

        store_position = store_position+b.num_of_clerks
        # print("pos ",store_position)

        # print(b.clerks)

    wb.save("data/"+str(StoreInfo.name)+".xlsx")

creat_target_file(myExcelRead.test_me())
